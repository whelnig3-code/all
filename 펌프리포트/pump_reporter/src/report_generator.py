"""v4.5: 월간 운영 리포트 생성기.

log_analyzer 결과를 기반으로 JSON + Excel 리포트 자동 생성.
출력: data/monthly_reports/monthly_report_YYYYMM.{json,xlsx}
"""
import json
import calendar
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.paths import DATA_DIR
from src.log_analyzer import analyze_decision_logs, analyze_fp_episodes
from src.logger import system_logger

logger = logging.getLogger(__name__)

MONTHLY_REPORTS_DIR = DATA_DIR / "monthly_reports"


# ── 운영 안정성 점수 ──────────────────────────────────────────

def _calculate_stability_score(summary: dict, fp_result: dict) -> int:
    """운영 안정성 점수 (0~100) 산출.

    산식 (4축 가중평균):
      (1) 평균 Confidence   : min(conf/100, 1.0) × 30
      (2) Log Coverage      : min(cov/100, 1.0) × 25
      (3) Rolling 경고 역가중: (1 - rolling_pct/100) × 25
      (4) FP Episode 역가중 : max(0, 1 - avg_ep_len/10) × 20
    """
    conf = summary.get("avg_baseline_confidence") or 0
    log_cov = summary.get("log_integrity", {}).get("coverage_pct", 0)
    rolling_pct = summary.get("rolling_warning_pct", 0)
    avg_ep = fp_result.get("avg_episode_length", 0)

    s1 = min(conf / 100, 1.0) * 30
    s2 = min(log_cov / 100, 1.0) * 25
    s3 = max(0, 1 - rolling_pct / 100) * 25
    s4 = max(0, 1 - avg_ep / 10) * 20

    return round(s1 + s2 + s3 + s4)


# ── 다음 달 위험 예측 ─────────────────────────────────────────

def _predict_risk_outlook(by_pump: dict, summary: dict) -> dict:
    """다음 달 위험 전망.

    고위험 펌프 조건 (하나라도 해당):
      - 월말 기준 forecast ≤ 6개월
      - rolling 경고 비율 50% 이상 (펌프 기준)
      - 평균 열화율 ≤ -15%
    """
    high_risk = []
    bases = []

    for pid, ps in by_pump.items():
        reasons = []
        fc = ps.get("avg_forecast_months")
        if fc is not None and fc <= 6:
            reasons.append(f"forecast {fc:.0f}개월")
        deg = ps.get("avg_degradation_pct")
        if deg is not None and deg <= -15:
            reasons.append(f"열화율 {deg:+.1f}%")
        # 경고일 비율
        rec_count = ps.get("record_count", 1)
        warn = ps.get("warning_days", 0)
        if rec_count > 0 and warn / rec_count >= 0.5:
            reasons.append(f"경고 {warn}/{rec_count}일")
        if reasons:
            high_risk.append(pid)
            bases.append(f"{pid}: {', '.join(reasons)}")

    return {
        "high_risk_pumps": high_risk,
        "high_risk_count": len(high_risk),
        "basis": bases,
    }


def generate_monthly_report(year: int, month: int) -> dict:
    """월간 운영 리포트 생성.

    Args:
        year: 연도 (e.g. 2026)
        month: 월 (1~12)

    Returns:
        {"json_path": Path, "xlsx_path": Path, "report": dict}
        로그 없으면 {"json_path": None, "xlsx_path": None, "report": None}
    """
    MONTHLY_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    start_date = f"{year}-{month:02d}-01"
    _, last_day = calendar.monthrange(year, month)
    end_date = f"{year}-{month:02d}-{last_day:02d}"

    # ── 로그 분석 ─────────────────────────────────────────────
    analysis = analyze_decision_logs(start_date, end_date)
    fp_result = analyze_fp_episodes(start_date, end_date)

    if analysis["summary"]["total_records"] == 0:
        logger.info(f"로그 없음: {year}-{month:02d}")
        return {"json_path": None, "xlsx_path": None, "report": None}

    summary = analysis["summary"]
    by_pump = analysis["by_pump"]

    # ── 안정성 점수 + 위험 예측 ───────────────────────────────
    stability_score = _calculate_stability_score(summary, fp_result)
    risk_outlook = _predict_risk_outlook(by_pump, summary)

    # ── 리포트 구조 생성 ──────────────────────────────────────
    report = {
        "generated_at": datetime.now().isoformat(),
        "period": f"{year}-{month:02d}",
        "stability_score": stability_score,
        "overview": {
            "total_analyses": summary["total_records"],
            "immediate_inspection_count": summary["final_category_distribution"].get("즉시점검", 0),
            "replacement_plan_count": summary["final_category_distribution"].get("교체계획", 0),
            "preventive_count": summary["final_category_distribution"].get("예방정비", 0),
            "normal_count": summary["final_category_distribution"].get("정상", 0),
            "fp_episode_count": fp_result["total_episodes"],
            "avg_baseline_confidence": summary["avg_baseline_confidence"],
            "avg_recent_coverage": summary["avg_recent_coverage"],
        },
        "log_integrity": summary["log_integrity"],
        "confidence_quality": {
            "low_confidence_pct": summary["low_confidence_pct"],
            "low_conf_immediate_count": summary["low_conf_immediate_count"],
        },
        "pump_performance": {},
        "stability_metrics": {
            "rolling_warning_pct": summary["rolling_warning_pct"],
            "forecast_usage_pct": summary["forecast_usage_pct"],
            "deg_severe_count": summary["deg_severe_count"],
            "avg_fp_episode_length": fp_result["avg_episode_length"],
            "max_fp_episode_length": fp_result["max_episode_length"],
        },
        "risk_outlook_next_month": risk_outlook,
        "category_distribution": summary["final_category_distribution"],
        "reason_distribution": summary["reason_distribution"],
    }

    # 펌프별 성능
    for pid, ps in by_pump.items():
        report["pump_performance"][pid] = {
            "avg_degradation_pct": ps["avg_degradation_pct"],
            "max_degradation_pct": ps["max_degradation_pct"],
            "avg_forecast_months": ps["avg_forecast_months"],
            "warning_days": ps["warning_days"],
            "record_count": ps["record_count"],
            "log_days": ps["log_days"],
            "log_completeness_pct": ps["log_completeness_pct"],
        }

    # ── JSON 저장 ─────────────────────────────────────────────
    suffix = f"{year}{month:02d}"
    json_path = MONTHLY_REPORTS_DIR / f"monthly_report_{suffix}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    # ── Excel 저장 ────────────────────────────────────────────
    xlsx_path = MONTHLY_REPORTS_DIR / f"monthly_report_{suffix}.xlsx"
    _generate_xlsx(report, analysis, fp_result, xlsx_path)

    try:
        system_logger.info({
            "timestamp": datetime.now().isoformat(),
            "event": "monthly_report_generated",
            "period": f"{year}-{month:02d}",
            "stability_score": stability_score,
            "high_risk_pumps": risk_outlook["high_risk_count"],
            "json_path": str(json_path),
            "xlsx_path": str(xlsx_path),
        })
    except Exception:
        pass

    return {"json_path": json_path, "xlsx_path": xlsx_path, "report": report}


# ── Excel 생성 ────────────────────────────────────────────────

_HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
_SCORE_FONT = Font(name="맑은 고딕", size=18, bold=True)


def _style_header(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _score_color(score: int) -> str:
    if score >= 80:
        return "2E7D32"
    if score >= 60:
        return "F57F17"
    return "C62828"


def _generate_xlsx(report: dict, analysis: dict,
                   fp_result: dict, path: Path):
    wb = Workbook()

    # ── Sheet 1: 종합요약 ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "종합요약"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 22

    ws1.cell(row=1, column=1, value=f"월간 운영 리포트 — {report['period']}").font = _TITLE_FONT
    ws1.merge_cells("A1:B1")

    # 안정성 점수
    score = report["stability_score"]
    ws1.cell(row=2, column=1, value="운영 안정성 점수").font = Font(
        name="맑은 고딕", size=11, bold=True)
    score_cell = ws1.cell(row=2, column=2, value=score)
    score_cell.font = Font(name="맑은 고딕", size=18, bold=True,
                           color=_score_color(score))
    score_cell.alignment = Alignment(horizontal="right")

    ov = report["overview"]
    li = report["log_integrity"]
    cq = report["confidence_quality"]
    sm = report["stability_metrics"]
    ro = report["risk_outlook_next_month"]

    rows = [
        ("", ""),
        ("[운영 개요]", ""),
        ("전체 분석 건수", ov["total_analyses"]),
        ("즉시점검 발생 횟수", ov["immediate_inspection_count"]),
        ("교체계획 발생 횟수", ov["replacement_plan_count"]),
        ("예방정비 발생 횟수", ov["preventive_count"]),
        ("정상 건수", ov["normal_count"]),
        ("오탐 에피소드 수", ov["fp_episode_count"]),
        ("평균 Baseline Confidence", ov["avg_baseline_confidence"]),
        ("평균 Recent Coverage (%)", ov["avg_recent_coverage"]),
        ("", ""),
        ("[로그 무결성]", ""),
        ("기대 일수", li["expected_days"]),
        ("실제 기록 일수", li["actual_days_logged"]),
        ("누락 일수", li["missing_days"]),
        ("로그 커버리지 (%)", li["coverage_pct"]),
        ("", ""),
        ("[Confidence 품질]", ""),
        ("낮은 Confidence 비율 (%)", cq["low_confidence_pct"]),
        ("저신뢰+즉시점검 건수", cq["low_conf_immediate_count"]),
        ("", ""),
        ("[안정성 지표]", ""),
        ("Rolling 기반 경고 비율 (%)", sm["rolling_warning_pct"]),
        ("Forecast 사용 비율 (%)", sm["forecast_usage_pct"]),
        ("deg_severe 발생 횟수", sm["deg_severe_count"]),
        ("평균 FP Episode 길이 (일)", sm["avg_fp_episode_length"]),
        ("최대 FP Episode 길이 (일)", sm["max_fp_episode_length"]),
        ("", ""),
        ("[다음 달 위험 예측]", ""),
        ("고위험 펌프 수", ro["high_risk_count"]),
        ("고위험 펌프", ", ".join(ro["high_risk_pumps"]) or "없음"),
    ]
    # 고위험 근거
    for basis_line in ro.get("basis", []):
        rows.append(("  근거", basis_line))

    for i, (label, value) in enumerate(rows, start=4):
        c1 = ws1.cell(row=i, column=1, value=label)
        c1.border = _THIN_BORDER
        if label.startswith("["):
            c1.font = Font(name="맑은 고딕", size=10, bold=True)
        c2 = ws1.cell(row=i, column=2, value=value)
        c2.border = _THIN_BORDER
        c2.alignment = Alignment(horizontal="right")

    # ── Sheet 2: 펌프별 통계 ──────────────────────────────────
    ws2 = wb.create_sheet("펌프별 통계")
    headers2 = ["펌프 ID", "분석 건수", "평균 열화율(%)", "최대 열화율(%)",
                "교체예측 평균(월)", "경고 발생 일수",
                "로그 일수", "로그 완전성(%)"]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, 1, len(headers2))

    low_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    for r_idx, (pid, ps) in enumerate(sorted(report["pump_performance"].items()), start=2):
        ws2.cell(row=r_idx, column=1, value=pid).border = _THIN_BORDER
        ws2.cell(row=r_idx, column=2, value=ps["record_count"]).border = _THIN_BORDER
        ws2.cell(row=r_idx, column=3, value=ps["avg_degradation_pct"]).border = _THIN_BORDER
        ws2.cell(row=r_idx, column=4, value=ps["max_degradation_pct"]).border = _THIN_BORDER
        ws2.cell(row=r_idx, column=5, value=ps["avg_forecast_months"]).border = _THIN_BORDER
        ws2.cell(row=r_idx, column=6, value=ps["warning_days"]).border = _THIN_BORDER
        ws2.cell(row=r_idx, column=7, value=ps["log_days"]).border = _THIN_BORDER
        c8 = ws2.cell(row=r_idx, column=8, value=ps["log_completeness_pct"])
        c8.border = _THIN_BORDER
        # 완전성 80% 미만 강조
        if ps["log_completeness_pct"] < 80:
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=r_idx, column=col).fill = low_fill

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[chr(64 + col)].width = 18

    # ── Sheet 3: 경고 타임라인 ────────────────────────────────
    ws3 = wb.create_sheet("경고 타임라인")
    headers3 = ["펌프 ID", "Episode 시작", "Episode 종료", "지속 일수"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    _style_header(ws3, 1, len(headers3))

    warn_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    crit_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for r_idx, ep in enumerate(fp_result.get("episodes", []), start=2):
        fill = crit_fill if ep["length"] >= 3 else warn_fill
        ws3.cell(row=r_idx, column=1, value=ep["pump_id"]).border = _THIN_BORDER
        ws3.cell(row=r_idx, column=2, value=ep["start"]).border = _THIN_BORDER
        ws3.cell(row=r_idx, column=3, value=ep["end"]).border = _THIN_BORDER
        cell = ws3.cell(row=r_idx, column=4, value=ep["length"])
        cell.border = _THIN_BORDER
        for col in range(1, 5):
            ws3.cell(row=r_idx, column=col).fill = fill

    for col in range(1, len(headers3) + 1):
        ws3.column_dimensions[chr(64 + col)].width = 18

    wb.save(str(path))
