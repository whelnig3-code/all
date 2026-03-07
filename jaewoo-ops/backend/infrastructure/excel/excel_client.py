"""
Excel 클라이언트 — Phase 1 데이터 저장소
openpyxl 기반으로 JAEWOO_OPS_운영대장.xlsx 를 읽고 씁니다.
"""
import os
from contextlib import contextmanager
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# 시트 이름 상수
SHEET_MAINTENANCE = "정기예방정비"
SHEET_EMERGENCY = "긴급업무기록"
SHEET_EMPLOYEES = "직원마스터"
SHEET_EVALUATION = "월간평가"
SHEET_TASK_PREFIX = "과제_"  # 개인과제 시트는 "과제_이름" 형태


class ExcelClient:
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)

    def _ensure_file(self) -> None:
        """Excel 파일이 없으면 템플릿으로 생성"""
        if not self.file_path.exists():
            self.file_path.parent.mkdir(parents=True, exist_ok=True)
            _create_template(self.file_path)

    @contextmanager
    def open_readonly(self):
        self._ensure_file()
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            yield wb
        finally:
            wb.close()

    @contextmanager
    def open_readwrite(self):
        """쓰기용 — 파일 잠금 위험이 있으므로 짧게 사용"""
        self._ensure_file()
        wb = openpyxl.load_workbook(self.file_path)
        try:
            yield wb
            wb.save(self.file_path)
        finally:
            wb.close()

    def read_sheet(self, sheet_name: str) -> list[dict[str, Any]]:
        """시트를 헤더-데이터 딕셔너리 리스트로 읽기"""
        with self.open_readonly() as wb:
            if sheet_name not in wb.sheetnames:
                return []
            ws: Worksheet = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h) if h is not None else f"col_{i}"
                       for i, h in enumerate(rows[0])]
            result = []
            for row in rows[1:]:
                if all(cell is None for cell in row):
                    continue
                result.append(dict(zip(headers, row)))
            return result

    def append_row(self, sheet_name: str, row_data: list[Any]) -> None:
        """시트에 행 추가"""
        with self.open_readwrite() as wb:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            ws.append(row_data)

    def update_row_by_id(self, sheet_name: str, id_col: str,
                          id_value: str, updates: dict[str, Any]) -> bool:
        """id_col 값이 일치하는 행을 업데이트. 성공 시 True 반환."""
        with self.open_readwrite() as wb:
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            try:
                id_col_idx = headers.index(id_col)
            except ValueError:
                return False

            for row in ws.iter_rows(min_row=2):
                if str(row[id_col_idx].value) == str(id_value):
                    for col_name, new_val in updates.items():
                        if col_name in headers:
                            col_idx = headers.index(col_name)
                            row[col_idx].value = new_val
                    return True
        return False

    def list_personal_task_sheets(self) -> list[str]:
        """개인과제 시트 목록 반환 (과제_xxx)"""
        with self.open_readonly() as wb:
            return [s for s in wb.sheetnames if s.startswith(SHEET_TASK_PREFIX)]


# ────────────────────────────────────────
# Excel 템플릿 생성
# ────────────────────────────────────────

def _create_template(path: Path) -> None:
    """JAEWOO_OPS_운영대장.xlsx 빈 템플릿 생성"""
    wb = Workbook()
    wb.remove(wb.active)

    # 시트 1: 정기예방정비
    ws1 = wb.create_sheet(SHEET_MAINTENANCE)
    ws1.append(["No", "점검항목", "설비명", "위치", "주기",
                "주담당자", "부담당자", "점검일", "마감일",
                "상태", "완료일시", "비고", "task_id"])

    # 시트 2: 긴급업무기록
    ws2 = wb.create_sheet(SHEET_EMERGENCY)
    ws2.append(["No", "긴급업무명", "발생일시", "대응시작", "대응종료",
                "소요시간", "대응자", "처리구분", "처리방법 상세",
                "비용/자재", "task_id"])

    # 시트 3: 직원마스터
    ws3 = wb.create_sheet(SHEET_EMPLOYEES)
    ws3.append(["사번", "이름", "연락처", "카카오워크ID",
                "팀", "직급", "언어", "지정백업", "상태",
                "이메일", "employee_id"])

    # 시트 4: 월간평가
    ws4 = wb.create_sheet(SHEET_EVALUATION)
    ws4.append(["월", "이름", "총배정", "정시완료", "지연완료",
                "미완료", "과제달성율", "설비점검가중치",
                "긴급대응평균시간", "종합점수", "등급", "employee_id"])

    wb.save(path)
    print(f"[ExcelClient] 템플릿 생성 완료: {path}")
