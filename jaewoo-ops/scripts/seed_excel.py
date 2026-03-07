"""
JAEWOO OPS — 샘플 데이터 생성 스크립트
Excel 운영대장에 직원, 설비, 업무 샘플 데이터를 추가합니다.
"""
import sys
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

EXCEL_PATH = Path(__file__).parent.parent / "excel" / "JAEWOO_OPS_운영대장.xlsx"


def seed():
    wb = load_workbook(EXCEL_PATH)

    # ── 1. 직원마스터 ─────────────────────────────────────────────────
    ws_emp = wb["직원마스터"]
    # 기존 데이터 확인 (헤더 제외)
    existing_employees = [row for row in ws_emp.iter_rows(min_row=2, values_only=True) if row[0]]
    if not existing_employees:
        employees = [
            # 사번, 이름, 연락처, 카카오워크ID, 팀, 직급, 언어, 지정백업, 상태, 이메일, employee_id
            ("E001", "김재우", "010-1234-5678", "kim.jaewoo", "재배팀", "팀장", "KO", "이민수", "재직", "jaewoo@company.com", str(uuid.uuid4())),
            ("E002", "이민수", "010-2345-6789", "lee.minsu", "재배팀", "사원", "KO", "박준혁", "재직", "minsu@company.com", str(uuid.uuid4())),
            ("E003", "박준혁", "010-3456-7890", "park.junhyuk", "재배팀", "사원", "KO", "김재우", "재직", "junhyuk@company.com", str(uuid.uuid4())),
            ("E004", "Nguyen Van An", "010-4567-8901", "nguyen.vanan", "재배팀", "사원", "VN", "이민수", "재직", "nguyen@company.com", str(uuid.uuid4())),
            ("E005", "관리자", "010-5678-9012", "admin.jaewoo", "관리팀", "관리자", "KO", "김재우", "재직", "admin@company.com", str(uuid.uuid4())),
        ]
        for emp in employees:
            ws_emp.append(emp)
        print(f"[Seed] 직원 {len(employees)}명 추가 완료")
    else:
        print(f"[Seed] 직원 데이터 이미 존재 ({len(existing_employees)}명), 스킵")

    # 직원 ID 매핑 (이름 → employee_id)
    emp_map = {}
    for row in ws_emp.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 사번 있으면
            emp_map[row[1]] = row[10]  # 이름 → employee_id

    # ── 2. 정기예방정비 ────────────────────────────────────────────────
    ws_maint = wb["정기예방정비"]
    existing_maint = [row for row in ws_maint.iter_rows(min_row=2, values_only=True) if row[0]]
    today = date.today()
    if not existing_maint:
        maint_tasks = [
            # No, 점검항목, 설비명, 위치, 주기, 주담당자, 부담당자, 점검일, 마감일, 상태, 완료일시, 비고, task_id
            ("001", "보일러 점검 및 압력 확인", "BL-001", "A동",  "주간", "김재우", "이민수",    "월요일",   str(today + timedelta(days=0)),  "대기", "", "", str(uuid.uuid4())),
            ("002", "냉동기 냉매 상태 점검",    "RF-003", "B동",  "월간", "이민수", "박준혁",    "매월 1일", str(today + timedelta(days=2)),  "대기", "", "냉매 부족 징후 확인 필요", str(uuid.uuid4())),
            ("003", "양액기 농도 측정",          "NS-002", "C동",  "일일", "박준혁", "이민수",    "매일",     str(today),                      "대기", "", "", str(uuid.uuid4())),
            ("004", "환기팬 필터 청소",           "VF-001", "A동",  "주간", "이민수", "김재우",    "수요일",   str(today + timedelta(days=3)),  "대기", "", "", str(uuid.uuid4())),
            ("005", "급수 펌프 작동 상태 확인",  "WP-002", "B동",  "일일", "Nguyen Van An", "박준혁", "매일",  str(today),                      "대기", "", "", str(uuid.uuid4())),
            ("006", "전기 패널 절연 저항 측정",  "EP-001", "관리동", "월간", "김재우", "관리자",  "매월 15일", str(today + timedelta(days=1)), "대기", "", "안전 최우선 작업", str(uuid.uuid4())),
            ("007", "CO2 공급기 압력 확인",       "CO2-001", "A동", "주간", "박준혁", "이민수",   "목요일",   str(today + timedelta(days=-1)), "지연", "", "", str(uuid.uuid4())),
            ("008", "하우스 온습도 센서 캘리브레이션", "TH-003", "C동", "월간", "이민수", "Nguyen Van An", "매월 20일", str(today + timedelta(days=5)), "대기", "", "", str(uuid.uuid4())),
        ]
        for task in maint_tasks:
            ws_maint.append(task)
        print(f"[Seed] 정기예방정비 {len(maint_tasks)}건 추가 완료")
    else:
        print(f"[Seed] 정기예방정비 데이터 이미 존재 ({len(existing_maint)}건), 스킵")

    # ── 3. 긴급업무기록 ────────────────────────────────────────────────
    ws_emg = wb["긴급업무기록"]
    existing_emg = [row for row in ws_emg.iter_rows(min_row=2, values_only=True) if row[0]]
    if not existing_emg:
        yesterday = today - timedelta(days=1)
        emg_tasks = [
            # No, 긴급업무명, 발생일시, 대응시작, 대응종료, 소요시간, 대응자, 처리구분, 처리방법 상세, 비용/자재, task_id
            (
                "001",
                "A동 보일러 배관 누수",
                str(yesterday) + " 14:30",
                str(yesterday) + " 14:45",
                str(yesterday) + " 16:20",
                "95분",
                "김재우",
                "자체처리",
                "배관 연결부 씰링 테이프 재시공 및 압력 테스트 완료",
                "씰링 테이프 3m, 공구 사용",
                str(uuid.uuid4()),
            ),
            (
                "002",
                "냉동기 이상 소음 발생",
                str(today) + " 09:15",
                str(today) + " 09:20",
                "",
                "진행중",
                "이민수",
                "외부업체",
                "냉동 전문 업체 긴급 출동 요청 (OO 냉동기술 010-1234-0000)",
                "출장비 예상 15만원",
                str(uuid.uuid4()),
            ),
        ]
        for task in emg_tasks:
            ws_emg.append(task)
        print(f"[Seed] 긴급업무기록 {len(emg_tasks)}건 추가 완료")
    else:
        print(f"[Seed] 긴급업무기록 데이터 이미 존재 ({len(existing_emg)}건), 스킵")

    # ── 4. 개인과제 시트 ───────────────────────────────────────────────
    personal_tasks = {
        "과제_김재우": [
            # No, 과제명, 카테고리, 배정일, 마감일, 상태, 완료일, 지연일수, 달성율, 평가메모, task_id
            ("001", "재배사 온도 자동화 개선안 작성", "개선업무", "2026-02-01", "2026-03-15", "진행", "", "", "60%", "", str(uuid.uuid4())),
            ("002", "2월 주간보고 작성", "보고", "2026-02-17", "2026-02-21", "완료", "2026-02-20", "0", "100%", "기한 내 완료", str(uuid.uuid4())),
            ("003", "신입 직원 OJT 자료 정리", "과제", "2026-02-10", "2026-02-28", "진행", "", "", "80%", "", str(uuid.uuid4())),
        ],
        "과제_이민수": [
            ("001", "설비 점검 매뉴얼 업데이트", "과제", "2026-02-05", "2026-02-28", "진행", "", "", "45%", "", str(uuid.uuid4())),
            ("002", "3월 농자재 발주 계획", "보고", "2026-02-18", "2026-02-25", "지연", "", "2", "30%", "외부 의존성으로 지연", str(uuid.uuid4())),
        ],
        "과제_박준혁": [
            ("001", "출하 프로세스 개선 보고서", "개선업무", "2026-01-20", "2026-02-29", "지연", "", "5", "40%", "데이터 수집 중", str(uuid.uuid4())),
            ("002", "3월 작물 생장 모니터링 계획", "과제", "2026-02-20", "2026-03-05", "진행", "", "", "20%", "", str(uuid.uuid4())),
        ],
    }

    for sheet_name, tasks in personal_tasks.items():
        if sheet_name not in wb.sheetnames:
            ws_p = wb.create_sheet(sheet_name)
            ws_p.append(["No", "과제명", "카테고리", "배정일", "마감일", "상태",
                         "완료일", "지연일수", "달성율", "평가메모", "task_id"])
            for t in tasks:
                ws_p.append(t)
            print(f"[Seed] {sheet_name} 시트 {len(tasks)}건 추가 완료")
        else:
            existing = [r for r in wb[sheet_name].iter_rows(min_row=2, values_only=True) if r[0]]
            if not existing:
                for t in tasks:
                    wb[sheet_name].append(t)
                print(f"[Seed] {sheet_name} {len(tasks)}건 추가 완료")
            else:
                print(f"[Seed] {sheet_name} 데이터 이미 존재, 스킵")

    # ── 5. 월간평가 샘플 ──────────────────────────────────────────────
    ws_eval = wb["월간평가"]
    existing_eval = [row for row in ws_eval.iter_rows(min_row=2, values_only=True) if row[0]]
    if not existing_eval:
        eval_data = [
            # 월, 이름, 총배정, 정시완료, 지연완료, 미완료, 과제달성율, 설비점검가중치, 긴급대응평균시간, 종합점수, 등급, employee_id
            ("2026-01", "김재우", 15, 12, 2, 1, "82%", "95점", "1.5시간", 88.5, "A", emp_map.get("김재우", "")),
            ("2026-01", "이민수", 12, 10, 1, 1, "75%", "88점", "2.0시간", 82.0, "A", emp_map.get("이민수", "")),
            ("2026-01", "박준혁", 10, 8,  2, 0, "70%", "85점", "없음",   79.5, "B", emp_map.get("박준혁", "")),
            ("2026-01", "Nguyen Van An", 8, 7, 1, 0, "88%", "90점", "없음", 85.0, "A", emp_map.get("Nguyen Van An", "")),
        ]
        for row in eval_data:
            ws_eval.append(row)
        print(f"[Seed] 월간평가 {len(eval_data)}건 추가 완료")
    else:
        print(f"[Seed] 월간평가 데이터 이미 존재, 스킵")

    wb.save(EXCEL_PATH)
    print(f"\n✅ 샘플 데이터 생성 완료: {EXCEL_PATH}")
    print("   시트 목록:", wb.sheetnames)


if __name__ == "__main__":
    seed()
