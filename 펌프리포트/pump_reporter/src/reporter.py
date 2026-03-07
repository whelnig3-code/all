"""Excel 리포트 생성기 – 통합 기간별 리포트.

기능:
  [1] 주간/월간/분기/연간 리포트 자동 생성
  [2] 요약 시트 + 전기 비교 시트 + 펌프별 상세 시트
  [3] 모든 펌프 리포트 생성 보장 (데이터부족 포함)
  [4] 파일명 규칙: {site}_펌프리포트_{type_kr}_{start}_to_{end}.xlsx
  [5] safe_save_workbook: PermissionError 방어
  [공통] ci.png 로고 삽입, 날짜 기반 표기 (Wxx 절대 미사용)
"""
import logging
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import font_manager
from calendar import monthrange

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

from src.config import (
    REPORTS_DIR, WEEKLY_DIR, MONTHLY_DIR, QUARTERLY_DIR, YEARLY_DIR,
    CHARTS_DIR, LOGO_PATH, load_settings,
)
from src.database import (
    get_all_pumps, get_casing_history, get_latest_baseline,
    get_daily_averages,
)
from src.analyzer import analyze_all_pumps, get_pump_trend_data
from src.period import (
    PERIOD_TYPE_KR, build_periods_with_data, prev_period,
    compare_periods, get_report_types,
)

logger = logging.getLogger(__name__)

# ── 한글 폰트 설정 ──────────────────────────────────────────
_FONT_CANDIDATES = ["Malgun Gothic", "맑은 고딕", "NanumGothic", "gulim"]
for _fn in _FONT_CANDIDATES:
    if _fn in [f.name for f in font_manager.fontManager.ttflist]:
        plt.rcParams["font.family"] = _fn
        break
plt.rcParams["axes.unicode_minus"] = False

# ── 스타일 상수 ─────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
NORMAL_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
JUDGMENT_COLORS = {
    "정밀점검": PatternFill("solid", fgColor="FF4444"),
    "점검권장": PatternFill("solid", fgColor="FFA500"),
    "경과관찰": PatternFill("solid", fgColor="FFFF00"),
    "정상": PatternFill("solid", fgColor="90EE90"),
    "데이터부족": PatternFill("solid", fgColor="D3D3D3"),
    "데이터없음": PatternFill("solid", fgColor="D3D3D3"),
    "분석오류": PatternFill("solid", fgColor="D3D3D3"),
    "기준 없음": PatternFill("solid", fgColor="E0E0FF"),
}
CHANGE_COLORS = {
    "개선": PatternFill("solid", fgColor="90EE90"),
    "악화": PatternFill("solid", fgColor="FF4444"),
    "안정": PatternFill("solid", fgColor="D3D3D3"),
    "전기 데이터 없음": PatternFill("solid", fgColor="EEEEEE"),
}
LOGO_HEIGHT_ROWS = 4

# matplotlib Figure 재사용 캐시 (렉 개선)
_REUSABLE_FIGURES: dict = {}


def _get_or_create_figure(key: str, figsize=(10, 5)):
    """matplotlib Figure 객체를 캐시하여 재사용. 메모리 절약."""
    if key not in _REUSABLE_FIGURES:
        fig, ax = plt.subplots(figsize=figsize)
        _REUSABLE_FIGURES[key] = (fig, ax)
    else:
        fig, ax = _REUSABLE_FIGURES[key]
        ax.clear()
        # 범례 등 초기화
        if fig.legends:
            for leg in fig.legends:
                leg.remove()
    return fig, ax

# 기간 유형별 출력 디렉토리
PERIOD_DIR_MAP = {
    "weekly": WEEKLY_DIR,
    "monthly": MONTHLY_DIR,
    "quarterly": QUARTERLY_DIR,
    "yearly": YEARLY_DIR,
}

# 추이 차트에서 보여줄 이전 기간 수
TREND_COUNTS = {
    "weekly": 4,
    "monthly": 4,
    "quarterly": 4,
    "yearly": 3,
}


# ═════════════════════════════════════════════════════════════
#  safe_save
# ═════════════════════════════════════════════════════════════
def safe_save_workbook(wb: Workbook, save_path: Path) -> Path:
    """저장 실패(파일 잠김) 시 timestamp 붙여 재시도. 최종 경로 반환."""
    save_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(save_path))
        return save_path
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = save_path.parent / f"{save_path.stem}_{ts}{save_path.suffix}"
        try:
            wb.save(str(fallback))
            logger.warning(f"원본 경로 잠김 → 대체 저장: {fallback}")
            return fallback
        except Exception as e:
            logger.error(f"저장 완전 실패: {e}")
            raise


# ═════════════════════════════════════════════════════════════
#  파일명
# ═════════════════════════════════════════════════════════════
def generate_report_filename(site: str, period_type: str,
                              p_start: str, p_end: str) -> str:
    """파일명: JAEWOO_펌프리포트_{type_kr}_{start}_to_{end}.xlsx"""
    type_kr = PERIOD_TYPE_KR.get(period_type, period_type)
    return f"JAEWOO_펌프리포트_{type_kr}_{p_start}_to_{p_end}.xlsx"


# ═════════════════════════════════════════════════════════════
#  공개 API - 통합 리포트 생성
# ═════════════════════════════════════════════════════════════
def generate_single_report(
        site: str, period_type: str,
        p_start: str, p_end: str,
        results: list[dict],
        prev_results: list[dict] | None,
        out_dir: Path = None,
) -> tuple[Path, bool]:
    """단일 기간 리포트 생성.

    Returns: (saved_path, success_bool)
    """
    if out_dir is None:
        out_dir = PERIOD_DIR_MAP.get(period_type, REPORTS_DIR)

    comparisons = compare_periods(results, prev_results, period_type)

    type_kr = PERIOD_TYPE_KR.get(period_type, period_type)
    period_str = f"{p_start} ~ {p_end}"

    # 전기 기간 정보
    if prev_results:
        ps, pe = prev_period(p_start, p_end, period_type)
    else:
        ps, pe = "", ""

    # 차트 생성
    comp_chart = _create_comparison_chart(
        results, prev_results or [],
        f"금기 ({p_start}~{p_end})",
        f"전기 ({ps}~{pe})" if ps else "전기",
        suffix=f"_{period_type}_{p_start}",
    )
    trend_chart = _create_multi_period_trend_chart(
        [r["pump_id"] for r in results],
        p_start, period_type,
    )

    # 워크북 생성
    wb = Workbook()

    _create_summary_sheet_with_comparison(
        wb, results, comparisons, site, period_str, type_kr,
        comp_chart, trend_chart)

    _create_comparison_detail_sheet(
        wb, comparisons, p_start, p_end, ps, pe, type_kr)

    for r in results:
        _create_pump_detail_sheet(wb, r, site, period_str)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # 저장
    fname = generate_report_filename(site, period_type, p_start, p_end)
    fpath = out_dir / fname

    try:
        actual_path = safe_save_workbook(wb, fpath)
        logger.info(f"{type_kr} 리포트 생성: {actual_path}")
        return actual_path, True
    except Exception as e:
        logger.error(f"{type_kr} 리포트 저장 실패: {e}")
        return fpath, False


def generate_all_reports(
        site: str, start: str, end: str,
        progress_callback=None,
) -> list[dict]:
    """사용자 선택 기간에 대해 데이터가 존재하는 기간만 리포트 일괄 생성.

    Returns: list of {"path", "saved_ok", "period_type", "period_type_kr",
                       "start", "end", "label"}
    """
    report_types = get_report_types(start, end)
    all_results = []

    # 데이터 있는 기간만 수집
    type_periods = {}
    for rt in report_types:
        periods = build_periods_with_data(start, end, rt)
        if periods:
            type_periods[rt] = periods

    if not type_periods:
        if progress_callback:
            progress_callback("생성 가능한 리포트가 없습니다.")
        return []

    total_count = sum(len(p) for p in type_periods.values())
    current = 0

    for rt, periods in type_periods.items():
        type_kr = PERIOD_TYPE_KR[rt]

        for p_start, p_end in periods:
            current += 1
            label = f"{type_kr} {p_start}~{p_end}"
            if progress_callback:
                progress_callback(
                    f"리포트 생성 중: {label} ({current}/{total_count})")

            try:
                this_results = analyze_all_pumps(p_start, p_end, rt)

                ps, pe = prev_period(p_start, p_end, rt)
                prev_results = []
                try:
                    prev_results = analyze_all_pumps(ps, pe, rt)
                except Exception:
                    prev_results = []

                path, ok = generate_single_report(
                    site, rt, p_start, p_end,
                    this_results, prev_results or None)

                all_results.append({
                    "path": str(path),
                    "saved_ok": ok,
                    "period_type": rt,
                    "period_type_kr": type_kr,
                    "start": p_start,
                    "end": p_end,
                    "label": label,
                })
            except Exception as e:
                logger.error(f"{label} 리포트 생성 실패: {e}")
                all_results.append({
                    "path": "",
                    "saved_ok": False,
                    "period_type": rt,
                    "period_type_kr": type_kr,
                    "start": p_start,
                    "end": p_end,
                    "label": label,
                })

    return all_results


# ═════════════════════════════════════════════════════════════
#  Legacy API (AnalysisWorker 호환)
# ═════════════════════════════════════════════════════════════
def generate_report(analysis_results: list[dict],
                    filename: str = None) -> tuple[Path, bool]:
    """전체 기간 리포트 1개 생성 (기존 호환용).

    Returns: (saved_path, report_saved_ok)
    """
    settings = load_settings()
    site = settings.get("site_name", "안평리")

    starts = [r["period_start"] for r in analysis_results
              if r.get("period_start")]
    ends = [r["period_end"] for r in analysis_results
            if r.get("period_end")]
    p_start = min(starts) if starts else "unknown"
    p_end = max(ends) if ends else "unknown"

    if not filename:
        filename = f"JAEWOO_펌프리포트_{p_start}_to_{p_end}.xlsx"

    report_path = REPORTS_DIR / filename
    try:
        actual_path = _build_legacy_workbook(
            analysis_results, report_path, site, f"{p_start} ~ {p_end}")
        logger.info(f"전체 리포트 생성: {actual_path}")
        return actual_path, True
    except Exception as e:
        logger.error(f"전체 리포트 저장 실패: {e}")
        return report_path, False


def _build_legacy_workbook(results: list[dict], save_path: Path,
                           site: str, period_str: str) -> Path:
    """Legacy 요약 + 펌프별 상세 워크북."""
    wb = Workbook()
    _create_legacy_summary_sheet(wb, results, site, period_str)
    for r in results:
        _create_pump_detail_sheet(wb, r, site, period_str)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]
    return safe_save_workbook(wb, save_path)


# ═════════════════════════════════════════════════════════════
#  요약 시트 (비교 컬럼 포함)
# ═════════════════════════════════════════════════════════════
def _create_summary_sheet_with_comparison(
        wb: Workbook, results: list[dict],
        comparisons: list[dict],
        site: str, period_str: str, type_kr: str,
        comp_chart_path: Path | None,
        trend_chart_path: Path | None):
    """31열 요약 시트: 전기 비교 + v3.1/v3.2/v3.3 컬럼 포함."""
    ws = wb.active
    ws.title = "요약"
    ncols = 31

    cur_row = _insert_logo(ws, "A1")

    # 제목
    title = f"{site} 지하수 펌프 유량 분석 보고서"
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=ncols)
    cell = ws.cell(row=cur_row, column=1, value=title)
    cell.font = Font(name="맑은 고딕", bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")
    cur_row += 1

    # 기간
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=ncols)
    ws.cell(row=cur_row, column=1,
            value=f"({period_str})").font = Font(
        name="맑은 고딕", size=10, color="555555")
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 1

    # 분석일
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=ncols)
    ws.cell(row=cur_row, column=1,
            value=f"분석일: {datetime.now().strftime('%Y-%m-%d')}").font = Font(
        name="맑은 고딕", size=9, color="888888")
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 1

    # 회사 내부용 안내
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=ncols)
    ws.cell(row=cur_row, column=1,
            value="재우 내부 보고서 시스템 v3.1 | 본 보고서는 ㈜재우 내부 운영 분석용 자료입니다. 외부 배포를 금합니다."
            ).font = Font(name="맑은 고딕", size=8, color="999999", italic=True)
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 2

    # 헤더
    headers = [
        "펌프ID", "분석일", "유효기간", "유효일수",
        "레코드(유효/기대)", "데이터 확보율(%)", "평균 유량(m³/h)",
        "기준 유량(m³/h)", "기준선 대비 변화(%)",
        "가동 패턴", "평균 가동시간(분/일)", "평균 가동횟수(회/일)", "주요 가동시간대",
        "판정", "전기 판정", "전기 대비", "전기 대비 유량변화(%)", "상세설명",
        "운전유형", "기준유량출처", "최고효율기준",
        "최고효율대비하락률(%)", "동일유형평균하락률(%)",
        "전체수량저하", "근거요약",
        "기준가동시간(분/일)", "가동시간출처", "가동시간감소율(%)",
        "반복점수", "미세OFF횟수", "미세OFF감지",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=cur_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    data_start = cur_row + 1

    comp_map = {c["pump_id"]: c for c in comparisons}

    for row_idx, r in enumerate(results, data_start):
        comp = comp_map.get(r.get("pump_id", ""), {})
        vr = r.get("valid_records", 0) or 0
        exp_r = r.get("expected_records", 0) or 0
        # effective 기간 우선, 없으면 valid 기간 fallback
        es = r.get("effective_start", "") or r.get("valid_start", "") or ""
        ee = r.get("effective_end", "") or r.get("valid_end", "") or ""
        ed = r.get("effective_days", 0) or r.get("valid_days", 0) or 0

        dr = r.get("data_rate")
        dr_display = f"{dr:.1f}" if dr is not None else "N/A"

        on_min = r.get("avg_on_minutes_per_day")
        on_evt = r.get("avg_on_events_per_day")
        be_bl = r.get("best_efficiency_baseline")
        be_deg = r.get("best_efficiency_degradation_pct")
        grp_deg = r.get("group_avg_degradation_pct")
        values = [
            r.get("pump_id", ""),
            r.get("analysis_date", ""),
            f"{es}~{ee}" if es else "",
            ed,
            f"{vr:,}/{exp_r:,}" if exp_r else f"{vr:,}/0",
            dr_display,
            r.get("avg_flow"),
            r.get("baseline_value"),
            r.get("degradation_pct"),
            r.get("timer_mode", ""),
            f"{on_min:.0f}" if on_min is not None else "",
            f"{on_evt:.1f}" if on_evt is not None else "",
            r.get("primary_on_window", ""),
            r.get("judgment", ""),
            comp.get("prev_judgment", "") or "전기 데이터 없음",
            comp.get("status_change", "전기 데이터 없음"),
            comp.get("flow_change_rate"),
            r.get("status_reason", ""),
            # v3.1 신규 7열
            r.get("operation_type", ""),
            r.get("baseline_source", ""),
            f"{be_bl:.2f}" if be_bl is not None else "",
            f"{be_deg:.1f}" if be_deg is not None else "",
            f"{grp_deg:.1f}" if grp_deg is not None else "",
            "O" if r.get("system_wide_drop") else "",
            _build_evidence_summary(r),
            # v3.2 가동시간 3열
            r.get("on_time_baseline"),
            r.get("on_time_baseline_source", ""),
            r.get("on_time_degradation_pct"),
            # v3.3 타이머 정밀화 3열
            r.get("timer_repeat_score"),
            r.get("micro_cycle_count", 0),
            "O" if r.get("micro_cycle_detected") else "",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # 판정 색상 (col 14)
        judgment = str(r.get("judgment", ""))
        jcell = ws.cell(row=row_idx, column=14)
        for key, fill in JUDGMENT_COLORS.items():
            if key in judgment:
                jcell.fill = fill
                break

        # 상태변화 색상 (col 16)
        change = comp.get("status_change", "전기 데이터 없음")
        ccell = ws.cell(row=row_idx, column=16)
        if change in CHANGE_COLORS:
            ccell.fill = CHANGE_COLORS[change]
        if change == "악화":
            ccell.font = Font(name="맑은 고딕", size=10, color="FFFFFF",
                              bold=True)

    last_data_row = data_start + len(results) - 1

    # 열 너비
    widths = [12, 10, 22, 7, 13, 8, 8, 8, 8,
              10, 8, 8, 10, 16, 16, 12, 10, 50,
              8, 10, 8, 10, 10, 8, 40,
              10, 10, 10, 8, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 차트 삽입
    chart_row = last_data_row + 3

    if comp_chart_path and comp_chart_path.exists():
        img = XlImage(str(comp_chart_path))
        img.width = 800
        img.height = 420
        ws.add_image(img, f"A{chart_row}")
        chart_row += 24

    if trend_chart_path and trend_chart_path.exists():
        img = XlImage(str(trend_chart_path))
        img.width = 800
        img.height = 420
        ws.add_image(img, f"A{chart_row}")
        chart_row += 24

    # Generated by PumpReporter 푸터
    ws.cell(row=chart_row + 1, column=1,
            value="Generated by PumpReporter").font = Font(
        name="맑은 고딕", size=8, color="999999", italic=True)


# ═════════════════════════════════════════════════════════════
#  전기 대비 비교 시트
# ═════════════════════════════════════════════════════════════
def _create_comparison_detail_sheet(
        wb: Workbook, comparisons: list[dict],
        this_start: str, this_end: str,
        prev_start: str, prev_end: str,
        type_kr: str):
    """전기 대비 상세 비교 시트."""
    ws = wb.create_sheet(title="전기비교")

    cur_row = 1
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=11)
    if prev_start:
        title = (f"전기 대비 비교  "
                 f"({this_start}~{this_end} vs "
                 f"{prev_start}~{prev_end})")
    else:
        title = f"전기 대비 비교  ({this_start}~{this_end})"
    ws.cell(row=cur_row, column=1, value=title).font = Font(
        name="맑은 고딕", bold=True, size=13)
    cur_row += 2

    headers = [
        "펌프ID", "금기 평균유량", "전기 평균유량", "유량 변화율(%)",
        "금기 확보율(%)", "전기 확보율(%)",
        "금기 기준선대비(%)", "전기 기준선대비(%)",
        "금기 판정", "전기 판정", "전기 대비",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=cur_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    cur_row += 1

    for comp in comparisons:
        this_dr = comp.get("this_data_rate")
        prev_dr = comp.get("prev_data_rate")
        values = [
            comp["pump_id"],
            comp.get("this_avg_flow"),
            comp.get("prev_avg_flow"),
            comp.get("flow_change_rate"),
            f"{this_dr:.1f}" if this_dr is not None else "N/A",
            f"{prev_dr:.1f}" if prev_dr is not None else "N/A",
            comp.get("this_degradation"),
            comp.get("prev_degradation"),
            comp.get("this_judgment", ""),
            comp.get("prev_judgment", "") or "전기 데이터 없음",
            comp.get("status_change", "전기 데이터 없음"),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=cur_row, column=col, value=v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        change = comp.get("status_change", "전기 데이터 없음")
        ccell = ws.cell(row=cur_row, column=11)
        if change in CHANGE_COLORS:
            ccell.fill = CHANGE_COLORS[change]
        if change == "악화":
            ccell.font = Font(name="맑은 고딕", size=10, color="FFFFFF",
                              bold=True)
        cur_row += 1

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14


# ═════════════════════════════════════════════════════════════
#  펌프별 상세 시트 — 데이터부족 펌프도 반드시 생성
# ═════════════════════════════════════════════════════════════
def _create_pump_detail_sheet(wb: Workbook, result: dict,
                              site: str, period_str: str):
    pump_id = result.get("pump_id", "unknown")
    sheet_name = pump_id[:28] if len(pump_id) > 28 else pump_id
    ws = wb.create_sheet(title=sheet_name)

    pumps = {p["pump_id"]: p for p in get_all_pumps()}
    pump_info = pumps.get(pump_id, {})

    cur_row = _insert_logo(ws, "A1")

    # 제목
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=6)
    ws.cell(row=cur_row, column=1,
            value=f"펌프 상세 분석: {pump_info.get('pump_name', pump_id)}"
            ).font = Font(name="맑은 고딕", bold=True, size=13)
    cur_row += 1

    ws.cell(row=cur_row, column=1,
            value=f"분석기간: {period_str}").font = Font(
        name="맑은 고딕", size=9, color="555555")
    cur_row += 1

    ws.cell(row=cur_row, column=1,
            value="재우 내부 보고서 시스템 v2.0 | 본 보고서는 ㈜재우 내부 운영 분석용 자료입니다. 외부 배포를 금합니다."
            ).font = Font(name="맑은 고딕", size=8, color="999999", italic=True)
    cur_row += 1

    # 기본 정보
    info_rows = [
        ("위치", pump_info.get("location", "")),
        ("설계용량 (m\u00b3/h)", pump_info.get("capacity_m3h", "")),
        ("설치일", pump_info.get("install_date", "")),
        ("점검주기 (일)", pump_info.get("inspect_cycle_days", "")),
    ]
    for label, val in info_rows:
        ws.cell(row=cur_row, column=1, value=label).font = Font(
            name="맑은 고딕", bold=True, size=10)
        ws.cell(row=cur_row, column=2, value=val).font = NORMAL_FONT
        cur_row += 1
    cur_row += 1

    # 분석 결과
    ws.cell(row=cur_row, column=1, value="분석 결과").font = Font(
        name="맑은 고딕", bold=True, size=11)
    cur_row += 1

    es = result.get("effective_start", "") or ""
    ee = result.get("effective_end", "") or ""
    ed = result.get("effective_days", 0) or 0
    exp_r = result.get("expected_records", 0) or 0
    vr = result.get("valid_records", 0) or 0

    dr = result.get("data_rate")
    dr_display = f"{dr:.1f}" if dr is not None else "N/A"

    analysis_rows = [
        ("분석일", result.get("analysis_date")),
        ("요청기간", f"{result.get('period_start', '')} ~ {result.get('period_end', '')}"),
        ("분석기간(effective)", f"{es} ~ {ee} ({ed}일)" if es else "데이터 없음"),
        ("기대 레코드", f"{exp_r:,}" if exp_r else "0"),
        ("유효 레코드", f"{vr:,}"),
        ("데이터 확보율 (%)", dr_display),
        ("평균 유량 (m\u00b3/h)", result.get("avg_flow")),
        ("최소 유량 (m\u00b3/h)", result.get("min_flow")),
        ("최대 유량 (m\u00b3/h)", result.get("max_flow")),
        ("기준 유량 (m\u00b3/h)", result.get("baseline_value")),
        ("기준선 대비 변화 (%)", result.get("degradation_pct")),
        ("가동 패턴", result.get("timer_mode", "") or "미분석"),
        ("평균 가동시간 (분/일)", result.get("avg_on_minutes_per_day")),
        ("평균 가동횟수 (회/일)", result.get("avg_on_events_per_day")),
        ("주요 가동시간대", result.get("primary_on_window", "") or "-"),
        ("마지막 교체 후 (일)", result.get("days_since_last_casing")),
        ("주기초과", "초과" if result.get("cycle_exceeded") else "정상"),
        ("종합 판정", result.get("judgment")),
        ("상세설명", result.get("status_reason", "")),
    ]
    for label, val in analysis_rows:
        cell_l = ws.cell(row=cur_row, column=1, value=label)
        cell_l.font = Font(name="맑은 고딕", bold=True, size=10)
        cell_l.border = THIN_BORDER
        cell_v = ws.cell(row=cur_row, column=2, value=val)
        cell_v.font = NORMAL_FONT
        cell_v.border = THIN_BORDER
        cell_v.alignment = Alignment(wrap_text=True)
        cur_row += 1

    # 판정 색상
    judgment = str(result.get("judgment", ""))
    j_cell = ws.cell(row=cur_row - 2, column=2)
    for key, fill in JUDGMENT_COLORS.items():
        if key in judgment:
            j_cell.fill = fill
            break
    cur_row += 1

    # 케이싱 이력
    casings = get_casing_history(pump_id)
    if casings:
        ws.cell(row=cur_row, column=1, value="케이싱 교체 이력").font = Font(
            name="맑은 고딕", bold=True, size=11)
        cur_row += 1
        ch_headers = ["교체일", "사유", "기준선초기화", "메모"]
        for col, h in enumerate(ch_headers, 1):
            cell = ws.cell(row=cur_row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
        cur_row += 1
        for ev in casings:
            vals = [ev["change_date"], ev.get("reason", ""),
                    "예" if ev.get("reset_baseline") else "아니오",
                    ev.get("memo", "")]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=cur_row, column=col, value=v)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
            cur_row += 1
    cur_row += 1

    # 차트
    chart_path = _create_pump_chart(pump_id, result)
    if chart_path:
        img = XlImage(str(chart_path))
        img.width = 750
        img.height = 400
        ws.add_image(img, f"A{cur_row}")

    # Generated by PumpReporter 푸터
    footer_row = cur_row + (24 if chart_path else 1)
    ws.cell(row=footer_row, column=1,
            value="Generated by PumpReporter").font = Font(
        name="맑은 고딕", size=8, color="999999", italic=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14


# ═════════════════════════════════════════════════════════════
#  차트 생성
# ═════════════════════════════════════════════════════════════
def _create_comparison_chart(
        this_results: list[dict], prev_results: list[dict],
        this_label: str, prev_label: str,
        suffix: str = "") -> Path | None:
    """금기 vs 전기 평균유량 비교 막대 차트."""
    fig, ax = _get_or_create_figure("comparison", (10, 5))

    pump_ids = [r["pump_id"] for r in this_results]
    if not pump_ids:
        plt.close(fig)
        return None

    prev_map = {r["pump_id"]: r for r in prev_results} if prev_results else {}

    this_flows = []
    prev_flows = []
    for r in this_results:
        this_flows.append(r.get("avg_flow") or 0)
        prev = prev_map.get(r["pump_id"])
        prev_flows.append(prev.get("avg_flow") or 0 if prev else 0)

    x = np.arange(len(pump_ids))
    width = 0.35

    ax.bar(x - width / 2, prev_flows, width, label=prev_label,
           color="#90CAF9", edgecolor="gray")
    ax.bar(x + width / 2, this_flows, width, label=this_label,
           color="#2196F3", edgecolor="gray")

    for bars in [ax.containers[0], ax.containers[1]]:
        for bar in bars:
            h = bar.get_height()
            if h > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, h,
                        f"{h:.1f}", ha="center", va="bottom", fontsize=8)

    ax.set_xticks(x)
    ax.set_xticklabels(pump_ids, rotation=30, ha="right")
    ax.set_title("금기 vs 전기 평균유량 비교", fontsize=13, fontweight="bold")
    ax.set_ylabel("평균유량 (m\u00b3/h)")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)

    fig.tight_layout()
    safe_suffix = suffix.replace("/", "_").replace("\\", "_")
    path = CHARTS_DIR / f"comparison{safe_suffix}.png"
    fig.savefig(str(path), dpi=150, bbox_inches="tight")
    plt.close(fig)
    return path


def _create_multi_period_trend_chart(
        pump_ids: list[str], ref_start: str,
        period_type: str) -> Path | None:
    """최근 N기간 추이 라인 차트."""
    fig, ax = _get_or_create_figure("trend", (10, 5))

    if not pump_ids:
        plt.close(fig)
        return None

    num_periods = TREND_COUNTS.get(period_type, 4)
    colors = ["#2196F3", "#FF5722", "#4CAF50", "#9C27B0", "#FF9800",
              "#795548", "#607D8B", "#E91E63"]

    # 기간 목록 생성 (현재 + 이전 N-1기간)
    ref_dt = datetime.strptime(ref_start, "%Y-%m-%d")

    periods = []
    cur_start = ref_start
    cur_end = ref_start  # 임시; 아래에서 계산

    # 현재 기간의 끝 결정
    if period_type == "weekly":
        weekday = ref_dt.weekday()
        mon = ref_dt - timedelta(days=weekday)
        sun = mon + timedelta(days=6)
        cur_start = mon.strftime("%Y-%m-%d")
        cur_end = sun.strftime("%Y-%m-%d")
    elif period_type == "monthly":
        _, last = monthrange(ref_dt.year, ref_dt.month)
        cur_start = ref_dt.replace(day=1).strftime("%Y-%m-%d")
        cur_end = ref_dt.replace(day=last).strftime("%Y-%m-%d")
    elif period_type == "quarterly":
        q_month = ((ref_dt.month - 1) // 3) * 3 + 1
        q_start = datetime(ref_dt.year, q_month, 1)
        q_end_month = q_month + 2
        _, last = monthrange(ref_dt.year, q_end_month)
        cur_start = q_start.strftime("%Y-%m-%d")
        cur_end = datetime(ref_dt.year, q_end_month, last).strftime("%Y-%m-%d")
    elif period_type == "yearly":
        cur_start = f"{ref_dt.year}-01-01"
        cur_end = f"{ref_dt.year}-12-31"

    # 현재부터 역순으로 N기간 수집
    p_s, p_e = cur_start, cur_end
    for _ in range(num_periods):
        periods.append((p_s, p_e))
        p_s, p_e = prev_period(p_s, p_e, period_type)
    periods.reverse()

    # 레이블 생성
    labels = []
    for ps, pe in periods:
        s = datetime.strptime(ps, "%Y-%m-%d")
        e = datetime.strptime(pe, "%Y-%m-%d")
        labels.append(f"{s.strftime('%m/%d')}~{e.strftime('%m/%d')}")

    for idx, pid in enumerate(pump_ids):
        values = []
        for ps, pe in periods:
            daily = get_daily_averages(pid, ps, pe)
            if daily:
                avg = np.mean([d["avg_flow"] for d in daily])
                values.append(round(avg, 2))
            else:
                values.append(0)

        color = colors[idx % len(colors)]
        ax.plot(labels, values, marker="o", linewidth=2,
                label=pid, color=color)

        for i, v in enumerate(values):
            if v > 0:
                ax.text(i, v, f"{v:.1f}", ha="center", va="bottom",
                        fontsize=7, color=color)

    type_kr = PERIOD_TYPE_KR.get(period_type, period_type)
    ax.set_title(f"최근 {num_periods}{type_kr} 평균유량 추이",
                 fontsize=13, fontweight="bold")
    ax.set_ylabel("평균유량 (m\u00b3/h)")
    ax.set_xlabel("기간")
    ax.legend(loc="best", fontsize=8)
    ax.grid(True, alpha=0.3)

    fig.tight_layout()
    path = CHARTS_DIR / f"trend_{period_type}_{ref_start}.png"
    fig.savefig(str(path), dpi=150, bbox_inches="tight")
    plt.close(fig)
    return path


def _create_pump_chart(pump_id: str, result: dict) -> Path | None:
    """펌프별 유량 추이 차트."""
    df = get_pump_trend_data(pump_id)

    fig, ax = _get_or_create_figure("pump_detail", (10, 5))

    if df.empty:
        ax.text(0.5, 0.5, "유효데이터가 없어 그래프를 표시할 수 없습니다.",
                ha="center", va="center", fontsize=14, color="gray",
                transform=ax.transAxes)
        ax.set_title(f"펌프 {pump_id} 유량 추이", fontsize=13, fontweight="bold")
        ax.set_xticks([])
        ax.set_yticks([])
    else:
        judgment = str(result.get("judgment", ""))
        is_insufficient = ("데이터부족" in judgment or "데이터없음" in judgment)

        ax.plot(df["date"], df["avg_flow"], marker="o", markersize=3,
                linewidth=1, label="일평균 유량", color="#2196F3", alpha=0.7)
        ax.plot(df["date"], df["ma7"], linewidth=2,
                label="7일 이동평균", color="#FF5722")

        baseline = get_latest_baseline(pump_id)
        if baseline:
            ax.axhline(y=baseline["baseline_value"], color="green",
                        linestyle="--", linewidth=1.5,
                        label=f"기준선 ({baseline['baseline_value']:.2f})")

        casings = get_casing_history(pump_id)
        for ev in casings:
            try:
                cd = pd.to_datetime(ev["change_date"])
                if df["date"].min() <= cd <= df["date"].max():
                    ax.axvline(x=cd, color="red", linestyle=":",
                               linewidth=1.5, alpha=0.8)
                    ax.annotate(f"교체({ev['change_date']})",
                                xy=(cd, ax.get_ylim()[1] * 0.95),
                                fontsize=7, color="red", rotation=45,
                                ha="left", va="top")
            except Exception:
                pass

        title = f"펌프 {pump_id} 유량 추이"
        if is_insufficient:
            title += " (유효데이터 부족)"
        ax.set_title(title, fontsize=13, fontweight="bold")
        ax.set_xlabel("날짜")
        ax.set_ylabel("유량 (m\u00b3/h)")
        ax.legend(loc="lower left", fontsize=8)
        ax.grid(True, alpha=0.3)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%m-%d"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator())
        fig.autofmt_xdate(rotation=45)

    fig.tight_layout()
    chart_path = CHARTS_DIR / f"chart_{pump_id}.png"
    fig.savefig(str(chart_path), dpi=150, bbox_inches="tight")
    plt.close(fig)
    return chart_path


def _create_summary_chart(results: list[dict]) -> Path | None:
    """요약 기준선 대비 변화 막대그래프."""
    fig, ax = _get_or_create_figure("summary", (10, 5))

    pump_ids = [r.get("pump_id", "?") for r in results]
    deg_vals = []
    for r in results:
        d = r.get("degradation_pct")
        deg_vals.append(d if d is not None else 0.0)

    if not pump_ids:
        ax.text(0.5, 0.5, "표시할 데이터 없음",
                ha="center", va="center", fontsize=16, color="gray",
                transform=ax.transAxes)
        ax.set_xticks([])
        ax.set_yticks([])
        ax.set_title("펌프별 기준선 대비 변화(%) 비교", fontsize=13, fontweight="bold")
    else:
        colors = []
        labels_below = []
        for d, r in zip(deg_vals, results):
            if r.get("degradation_pct") is None:
                colors.append("#BBBBBB")
                labels_below.append("N/A")
            elif d <= -20:
                colors.append("#FF4444")
                labels_below.append(f"{d:.1f}%")
            elif d <= -10:
                colors.append("#FFA500")
                labels_below.append(f"{d:.1f}%")
            elif d <= -5:
                colors.append("#FFD700")
                labels_below.append(f"{d:.1f}%")
            else:
                colors.append("#4CAF50")
                labels_below.append(f"{d:.1f}%")

        x_pos = range(len(pump_ids))
        bars = ax.bar(x_pos, deg_vals, color=colors, edgecolor="gray",
                      width=0.6)

        ax.axhline(y=0, color="black", linewidth=0.5)
        ax.axhline(y=-5, color="#FFD700", linestyle="--", linewidth=0.8,
                    alpha=0.5, label="경과관찰 (-5%)")
        ax.axhline(y=-10, color="#FFA500", linestyle="--", linewidth=0.8,
                    alpha=0.5, label="점검권장 (-10%)")
        ax.axhline(y=-20, color="#FF4444", linestyle="--", linewidth=0.8,
                    alpha=0.5, label="정밀점검 (-20%)")

        for bar, lbl in zip(bars, labels_below):
            y = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, y,
                    lbl, ha="center",
                    va="bottom" if y >= 0 else "top", fontsize=9,
                    fontweight="bold")

        y_min = min(-25, min(deg_vals) - 5 if deg_vals else -25)
        y_max = max(10, max(deg_vals) + 5 if deg_vals else 10)
        ax.set_ylim(y_min, y_max)

        ax.set_xticks(x_pos)
        ax.set_xticklabels(pump_ids, rotation=30, ha="right")
        ax.set_title("펌프별 기준선 대비 변화(%) 비교", fontsize=13, fontweight="bold")
        ax.set_ylabel("기준선 대비 변화 (%)")
        ax.legend(loc="lower right", fontsize=7)
        ax.grid(axis="y", alpha=0.3)

    fig.tight_layout()
    chart_path = CHARTS_DIR / "summary_chart.png"
    fig.savefig(str(chart_path), dpi=150, bbox_inches="tight")
    plt.close(fig)
    return chart_path


# ═════════════════════════════════════════════════════════════
#  v3.1 근거요약 헬퍼
# ═════════════════════════════════════════════════════════════
def _build_evidence_summary(r: dict) -> str:
    """판정 근거 요약 1줄 생성."""
    parts = []
    op = r.get("operation_type", "")
    if op:
        parts.append(f"역할={op}")
    bs = r.get("baseline_source", "")
    if bs:
        parts.append(f"기준={bs}")
    deg = r.get("degradation_pct")
    if deg is not None:
        parts.append(f"하락={deg:.1f}%")
    be_deg = r.get("best_efficiency_degradation_pct")
    if be_deg is not None:
        parts.append(f"최고효율대비={be_deg:.1f}%")
    on_deg = r.get("on_time_degradation_pct")
    if on_deg is not None:
        parts.append(f"가동시간={on_deg:.1f}%")
    if r.get("system_wide_drop"):
        parts.append("전체하락")
    score = r.get("timer_repeat_score")
    if score is not None:
        parts.append(f"반복={score:.2f}")
    if r.get("micro_cycle_detected"):
        parts.append(f"미세OFF={r.get('micro_cycle_count', 0)}회")
    return " | ".join(parts) if parts else ""


# ═════════════════════════════════════════════════════════════
#  Legacy 요약 시트 (AnalysisWorker용)
# ═════════════════════════════════════════════════════════════
def _create_legacy_summary_sheet(wb: Workbook, results: list[dict],
                                 site: str, period_str: str):
    ws = wb.active
    ws.title = "요약"
    ncols = 26

    cur_row = _insert_logo(ws, "A1")

    title = f"{site} 지하수 펌프 유량 분석 보고서"
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=ncols)
    cell = ws.cell(row=cur_row, column=1, value=title)
    cell.font = Font(name="맑은 고딕", bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")
    cur_row += 1

    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=ncols)
    ws.cell(row=cur_row, column=1,
            value=f"({period_str})").font = Font(
        name="맑은 고딕", size=10, color="555555")
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 1

    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=ncols)
    ws.cell(row=cur_row, column=1,
            value=f"분석일: {datetime.now().strftime('%Y-%m-%d')}").font = Font(
        name="맑은 고딕", size=9, color="888888")
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 1

    # 회사 내부용 안내
    ws.merge_cells(start_row=cur_row, start_column=1,
                   end_row=cur_row, end_column=ncols)
    ws.cell(row=cur_row, column=1,
            value="재우 내부 보고서 시스템 v3.1 | 본 보고서는 ㈜재우 내부 운영 분석용 자료입니다. 외부 배포를 금합니다."
            ).font = Font(name="맑은 고딕", size=8, color="999999", italic=True)
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center")
    cur_row += 2

    headers = [
        "펌프ID", "분석일", "전체기간", "유효기간", "유효일수",
        "레코드(유효/기대)", "데이터 확보율(%)", "평균 유량(m³/h)", "기준 유량(m³/h)",
        "기준선 대비 변화(%)", "가동 패턴", "판정", "상세설명",
        "운전유형", "기준유량출처", "최고효율기준",
        "최고효율대비하락률(%)", "동일유형평균하락률(%)",
        "전체수량저하", "근거요약",
        "기준가동시간(분/일)", "가동시간출처", "가동시간감소율(%)",
        "반복점수", "미세OFF횟수", "미세OFF감지",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=cur_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    data_start_row = cur_row + 1

    for row_idx, r in enumerate(results, data_start_row):
        vr = r.get("valid_records", 0) or 0
        exp_r = r.get("expected_records", 0) or 0
        es = r.get("effective_start", "") or r.get("valid_start", "") or ""
        ee = r.get("effective_end", "") or r.get("valid_end", "") or ""
        ed = r.get("effective_days", 0) or r.get("valid_days", 0) or 0

        dr = r.get("data_rate")
        dr_display = f"{dr:.1f}" if dr is not None else "N/A"

        be_bl = r.get("best_efficiency_baseline")
        be_deg = r.get("best_efficiency_degradation_pct")
        grp_deg = r.get("group_avg_degradation_pct")
        values = [
            r.get("pump_id", ""),
            r.get("analysis_date", ""),
            f"{r.get('period_start','')}~{r.get('period_end','')}",
            f"{es}~{ee}" if es else "",
            ed,
            f"{vr:,}/{exp_r:,}" if exp_r else f"{vr:,}/0",
            dr_display,
            r.get("avg_flow"),
            r.get("baseline_value"),
            r.get("degradation_pct"),
            r.get("timer_mode", ""),
            r.get("judgment", ""),
            r.get("status_reason", ""),
            # v3.1 신규 7열
            r.get("operation_type", ""),
            r.get("baseline_source", ""),
            f"{be_bl:.2f}" if be_bl is not None else "",
            f"{be_deg:.1f}" if be_deg is not None else "",
            f"{grp_deg:.1f}" if grp_deg is not None else "",
            "O" if r.get("system_wide_drop") else "",
            _build_evidence_summary(r),
            # v3.2 가동시간 3열
            r.get("on_time_baseline"),
            r.get("on_time_baseline_source", ""),
            r.get("on_time_degradation_pct"),
            # v3.3 타이머 정밀화 3열
            r.get("timer_repeat_score"),
            r.get("micro_cycle_count", 0),
            "O" if r.get("micro_cycle_detected") else "",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        judgment = str(r.get("judgment", ""))
        jcell = ws.cell(row=row_idx, column=12)
        for key, fill in JUDGMENT_COLORS.items():
            if key in judgment:
                jcell.fill = fill
                break

    last_data_row = data_start_row + len(results) - 1

    col_widths = [12, 10, 22, 22, 7, 13, 8, 8, 8, 8, 6, 16, 50,
                  8, 10, 8, 10, 10, 8, 40,
                  10, 10, 10, 8, 8, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    chart_row = last_data_row + 3
    chart_path = _create_summary_chart(results)
    if chart_path:
        img = XlImage(str(chart_path))
        img.width = 800
        img.height = 420
        ws.add_image(img, f"A{chart_row}")


# ═════════════════════════════════════════════════════════════
#  유틸리티
# ═════════════════════════════════════════════════════════════
def _insert_logo(ws, cell_ref: str) -> int:
    """로고 삽입. 성공 시 로고 아래 첫 사용 가능 행 반환."""
    try:
        if LOGO_PATH.exists():
            img = XlImage(str(LOGO_PATH))
            img.width = 160
            img.height = 60
            ws.add_image(img, cell_ref)
            return LOGO_HEIGHT_ROWS + 1
        return 1
    except Exception as e:
        logger.warning(f"로고 삽입 실패: {e}")
        return 1
